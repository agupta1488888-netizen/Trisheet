"""m12 — output assembly.

Responsibility
    Turn a verified report into the forms it is consumed in: the JSON document
    the browser renders, a PDF styled as an analyst trisheet, and an XLSX
    whose ratios are live formulas over the reported figures. Source references
    travel with every one of them — the provenance rail is not a screen-only
    feature.

One document, three renderings
    `assemble_document` builds the canonical `ReportDocument` once; the PDF and
    the workbook are rendered from that object rather than from the fact store.
    That is deliberate: a figure cannot appear in the PDF having failed
    verification on the screen, because both are printing the same assembled
    document.

Why the workbook carries formulas
    A reader who disagrees with a margin should be able to see the division
    rather than take it on trust. Every derived figure in the workbook is an
    Excel formula pointing at the reported cells it came from, so changing an
    input changes the result. The assumptions tab holds the tunables those
    formulas reference, which is what makes them assumptions rather than
    constants buried in code.

Public interface
    assemble_document(inputs) -> ReportDocument
    render_html(document) -> str
    assemble_pdf(document) -> bytes
    assemble_xlsx(document) -> bytes
    publish(report_id, document) -> tuple[ArtifactRef, ...]
"""

from __future__ import annotations

import datetime as dt
import html
import logging
from collections.abc import Callable, Iterable, Mapping, Sequence
from dataclasses import dataclass, field
from typing import TYPE_CHECKING

from app.config import (
    ARTIFACT_CONTENT_TYPES,
    DAYS_IN_YEAR,
    DISPLAY_SCALE_DIVISOR,
    DISPLAY_SCALE_NOTE,
    MAX_CHART_SEGMENTS,
    MAX_TABLE_PERIODS,
    NOT_DISCLOSED_TEXT,
    SEGMENT_METRIC,
    UNSCALED_METRIC_PREFIXES,
)
from app.models import (
    AnalysisDepth,
    ArtifactKind,
    ArtifactRef,
    CashFlowPoint,
    CashFlowSeries,
    Company,
    ComplianceReport,
    DevelopmentEvent,
    DocumentFact,
    DocumentSection,
    Fact,
    FigureEmphasis,
    FigureRow,
    FigureTable,
    FilingRef,
    GeneratedReport,
    PeerSet,
    PeerValuationPoint,
    PeerValuationSeries,
    ProseBlock,
    Report,
    ReportCharts,
    ReportDocument,
    RevenueMarginPoint,
    RevenueMarginSeries,
    RiskItem,
    SectionId,
    SegmentMixPoint,
    SegmentMixSeries,
    SeriesMeta,
    SourceNote,
)
from app.modules import m04_narrative, m08_peers
from app.services import storage

if TYPE_CHECKING:
    # Imported for annotations only. Neither renderer is a hard dependency:
    # importing this module must work on a machine where WeasyPrint's native
    # libraries are absent, so both are imported inside the functions that
    # render rather than at module scope.
    from openpyxl.styles import Alignment, Border, Font
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

#: What a worksheet cell may hold. openpyxl accepts more than this; the
#: workbook writes only these. None is included deliberately: a fact the filer
#: did not disclose leaves the numeric cell empty rather than writing a zero,
#: while its display value on the same row still reads "Not disclosed".
_CellValue = str | float | int | None

#: Resolves a metric reference to the workbook cell holding it, or None when
#: the filer did not report that figure for the period being built.
_CellRef = Callable[[str], str | None]

#: Builds one analysis cell. Returns None when an input it needs is missing,
#: which is how a formula declines to exist rather than dividing by a blank.
_FormulaBuilder = Callable[[_CellRef], str | None]


class ArtifactRenderError(Exception):
    """An artifact could not be rendered.

    Carries a message written for a reader, because it is shown in place of a
    download link rather than logged and forgotten.
    """


# --- Table definitions -------------------------------------------------------
# Which metrics appear in which table, and in what order. Kept here rather than
# in config because these are the assembler's own presentation choices, not
# tunables the rest of the system shares.


@dataclass(frozen=True, slots=True)
class _Row:
    """One row of a figure table."""

    metric: str
    label: str
    emphasis: FigureEmphasis | None = None


_INCOME_ROWS: tuple[_Row, ...] = (
    _Row("income.revenue", "Revenue", FigureEmphasis.TOTAL),
    _Row("income.cost_of_revenue", "Cost of revenue"),
    _Row("income.gross_profit", "Gross profit", FigureEmphasis.TOTAL),
    _Row("income.research_and_development", "Research and development"),
    _Row("income.selling_general_administrative", "Selling, general and admin"),
    _Row("income.operating_income", "Operating income", FigureEmphasis.TOTAL),
    _Row("income.interest_expense", "Interest expense"),
    _Row("income.pretax_income", "Income before tax"),
    _Row("income.income_tax_expense", "Income tax expense"),
    _Row("income.net_income", "Net income", FigureEmphasis.TOTAL),
    _Row("income.eps_diluted", "Diluted earnings per share"),
    _Row("income.shares_diluted", "Diluted weighted average shares"),
)

_BALANCE_ROWS: tuple[_Row, ...] = (
    _Row("balance.cash_and_equivalents", "Cash and cash equivalents"),
    _Row("balance.accounts_receivable", "Accounts receivable"),
    _Row("balance.inventory", "Inventory"),
    _Row("balance.current_assets", "Total current assets", FigureEmphasis.TOTAL),
    _Row("balance.total_assets", "Total assets", FigureEmphasis.TOTAL),
    _Row("balance.accounts_payable", "Accounts payable"),
    _Row("balance.current_liabilities", "Total current liabilities"),
    _Row("balance.short_term_debt", "Short-term debt"),
    _Row("balance.long_term_debt", "Long-term debt"),
    _Row("balance.total_liabilities", "Total liabilities", FigureEmphasis.TOTAL),
    _Row("balance.total_equity", "Total equity", FigureEmphasis.TOTAL),
)

_CASHFLOW_ROWS: tuple[_Row, ...] = (
    _Row("cashflow.operating", "Net cash from operating activities"),
    _Row("cashflow.investing", "Net cash from investing activities"),
    _Row("cashflow.financing", "Net cash from financing activities"),
    _Row("cashflow.capital_expenditure", "Capital expenditure"),
    _Row(
        "cashflow.free_cash_flow",
        "Free cash flow",
        FigureEmphasis.DERIVED,
    ),
    _Row("cashflow.dividends_paid", "Dividends paid"),
    _Row("cashflow.share_repurchases", "Share repurchases"),
)

_MARGIN_ROWS: tuple[_Row, ...] = (
    _Row("margin.gross", "Gross margin", FigureEmphasis.DERIVED),
    _Row("margin.operating", "Operating margin", FigureEmphasis.DERIVED),
    _Row("margin.net", "Net margin", FigureEmphasis.DERIVED),
    _Row("margin.ebitda", "EBITDA margin", FigureEmphasis.DERIVED),
    _Row("return.roe", "Return on equity", FigureEmphasis.DERIVED),
    _Row("return.roa", "Return on assets", FigureEmphasis.DERIVED),
    _Row("return.roic", "Return on invested capital", FigureEmphasis.DERIVED),
    _Row(
        "return.effective_tax_rate",
        "Effective tax rate",
        FigureEmphasis.DERIVED,
    ),
)

_STRENGTH_ROWS: tuple[_Row, ...] = (
    _Row("liquidity.current_ratio", "Current ratio", FigureEmphasis.DERIVED),
    _Row("liquidity.quick_ratio", "Quick ratio", FigureEmphasis.DERIVED),
    _Row("leverage.debt_to_equity", "Debt to equity", FigureEmphasis.DERIVED),
    _Row(
        "leverage.net_debt_to_ebitda",
        "Net debt to EBITDA",
        FigureEmphasis.DERIVED,
    ),
    _Row(
        "leverage.interest_coverage",
        "Interest coverage",
        FigureEmphasis.DERIVED,
    ),
    _Row(
        "cashflow.fcf_conversion",
        "Free cash flow conversion",
        FigureEmphasis.DERIVED,
    ),
    _Row(
        "working_capital.cash_conversion_cycle",
        "Cash conversion cycle",
        FigureEmphasis.DERIVED,
    ),
)

_GROWTH_ROWS: tuple[_Row, ...] = (
    _Row("growth.income.revenue.yoy", "Revenue growth", FigureEmphasis.DERIVED),
    _Row(
        "growth.income.gross_profit.yoy",
        "Gross profit growth",
        FigureEmphasis.DERIVED,
    ),
    _Row(
        "growth.income.operating_income.yoy",
        "Operating income growth",
        FigureEmphasis.DERIVED,
    ),
    _Row(
        "growth.income.net_income.yoy",
        "Net income growth",
        FigureEmphasis.DERIVED,
    ),
    _Row(
        "growth.income.eps_diluted.yoy",
        "Diluted EPS growth",
        FigureEmphasis.DERIVED,
    ),
    _Row(
        "growth.cashflow.operating.yoy",
        "Operating cash flow growth",
        FigureEmphasis.DERIVED,
    ),
)

_MARKET_ROWS: tuple[_Row, ...] = (
    _Row("market.price", "Price"),
    _Row("market.market_cap", "Market capitalisation"),
    _Row("market.previous_close", "Previous close"),
    _Row("market.fifty_two_week_high", "52-week high"),
    _Row("market.fifty_two_week_low", "52-week low"),
    _Row("market.volume", "Volume"),
)

#: Sector tables, appended after the general ones when the filer has them.
_SECTOR_ROWS: tuple[tuple[str, str, tuple[_Row, ...]], ...] = (
    (
        "bank",
        "Banking metrics",
        (
            _Row("bank.net_interest_income", "Net interest income"),
            _Row(
                "bank.net_interest_margin",
                "Net interest margin",
                FigureEmphasis.DERIVED,
            ),
            _Row(
                "bank.efficiency_ratio",
                "Efficiency ratio",
                FigureEmphasis.DERIVED,
            ),
            _Row("bank.total_loans", "Loans and leases"),
            _Row("bank.total_deposits", "Total deposits"),
            _Row(
                "bank.loan_to_deposit",
                "Loan to deposit",
                FigureEmphasis.DERIVED,
            ),
            _Row("bank.cet1_ratio", "CET1 ratio", FigureEmphasis.DERIVED),
        ),
    ),
    (
        "reit",
        "REIT metrics",
        (
            _Row("reit.ffo", "Funds from operations", FigureEmphasis.DERIVED),
            _Row(
                "reit.affo",
                "Adjusted funds from operations",
                FigureEmphasis.DERIVED,
            ),
            _Row("reit.ffo_per_share", "FFO per share", FigureEmphasis.DERIVED),
            _Row(
                "reit.ffo_payout_ratio",
                "FFO payout ratio",
                FigureEmphasis.DERIVED,
            ),
        ),
    ),
    (
        "insurance",
        "Underwriting metrics",
        (
            _Row("insurance.earned_premiums", "Net premiums earned"),
            _Row("insurance.loss_ratio", "Loss ratio", FigureEmphasis.DERIVED),
            _Row(
                "insurance.expense_ratio",
                "Expense ratio",
                FigureEmphasis.DERIVED,
            ),
            _Row(
                "insurance.combined_ratio",
                "Combined ratio",
                FigureEmphasis.DERIVED,
            ),
        ),
    ),
)

#: Section titles. Mirrors WRITER_SECTIONS but is the assembler's own, because
#: a section can be rendered with tables when no prose was written for it.
_SECTION_TITLES: dict[SectionId, str] = {
    SectionId.SNAPSHOT: "Snapshot",
    SectionId.BUSINESS: "Business",
    SectionId.FINANCIALS: "Financial highlights",
    SectionId.ANALYSIS: "Analysis",
    SectionId.PEERS: "Peers",
    SectionId.DEVELOPMENTS: "Recent developments",
    SectionId.RISKS: "Risks",
}

#: What an empty section says. Never a blank panel, and never "N/A".
_UNAVAILABLE: dict[SectionId, str] = {
    SectionId.SNAPSHOT: (
        "No market data was available for this filer, so there is no snapshot."
    ),
    SectionId.BUSINESS: (
        "The business description could not be read from the annual report."
    ),
    SectionId.FINANCIALS: (
        "No reported figures were extracted for this filer."
    ),
    SectionId.ANALYSIS: (
        "Not enough periods were reported to derive comparable metrics."
    ),
    SectionId.PEERS: (
        "No filed source named a comparable company for this filer."
    ),
    SectionId.DEVELOPMENTS: (
        "This filer has filed no current reports covering a reportable event."
    ),
    SectionId.RISKS: (
        "The risk factors item could not be read from the annual report."
    ),
}


@dataclass(frozen=True, slots=True)
class AssemblyInput:
    """Everything the assembler needs. One object, so the signature is stable.

    The report is assembled from what the pipeline already holds in memory; it
    is not re-read from the fact store. Re-reading would let the two diverge,
    and the assembled document is what was verified.
    """

    report: Report
    company: Company
    depth: AnalysisDepth
    facts: Sequence[Fact]
    filings: Sequence[FilingRef]
    compliance: ComplianceReport
    prose: GeneratedReport | None = None
    peers: PeerSet | None = None
    peer_comparison: m08_peers.PeerComparison | None = None
    events: Sequence[DevelopmentEvent] = ()
    artifacts: Sequence[ArtifactRef] = ()
    #: Statements read off links the reader supplied. Carried through to the
    #: document untouched and never indexed with `facts` — they are not facts,
    #: and `_build_index` is what feeds the tables.
    source_notes: Sequence[SourceNote] = ()


# --- Fact indexing -----------------------------------------------------------


@dataclass(slots=True)
class _Index:
    """Facts arranged for lookup by metric and period.

    Built once per document. Everything downstream — tables, charts, the
    workbook — reads through this, so a figure that appears in two places is
    the same fact in both.
    """

    by_metric_year: dict[tuple[str, int], Fact] = field(default_factory=dict)
    #: Latest fact per metric, for figures that have no fiscal year (market
    #: data, peers, narrative).
    latest: dict[str, Fact] = field(default_factory=dict)
    #: Segment revenue by member label then fiscal year. Held apart because a
    #: segment shares its metric with its siblings and cannot be addressed by
    #: metric alone.
    segments: dict[str, dict[int, Fact]] = field(default_factory=dict)
    years: tuple[int, ...] = ()
    currency: str | None = None

    def fact(self, metric: str, year: int) -> Fact | None:
        return self.by_metric_year.get((metric, year))

    def has(self, metric: str) -> bool:
        return metric in self.latest


def _build_index(facts: Sequence[Fact]) -> _Index:
    """Indexes facts by metric and fiscal year, newest wins on a tie."""
    index = _Index()

    for fact in facts:
        held = index.latest.get(fact.metric)
        if held is None or _is_later(fact, held):
            index.latest[fact.metric] = fact

        if fact.fiscal_year is None:
            continue
        key = (fact.metric, fact.fiscal_year)
        # A segment fact shares its metric with its siblings; it is addressed
        # by segment elsewhere and must not displace the consolidated figure.
        if fact.segment_member is not None:
            continue
        incumbent = index.by_metric_year.get(key)
        if incumbent is None or _is_later(fact, incumbent):
            index.by_metric_year[key] = fact

    index.segments = _segment_series(facts)
    index.years = _periods(index)
    index.currency = _currency(facts)
    return index


def _is_later(candidate: Fact, incumbent: Fact) -> bool:
    """Whether `candidate` supersedes `incumbent` for the same slot."""
    if candidate.filed_date != incumbent.filed_date:
        return candidate.filed_date > incumbent.filed_date
    return candidate.period_end > incumbent.period_end


def _periods(index: _Index) -> tuple[int, ...]:
    """Fiscal years shown, oldest first.

    Chosen from the years revenue and total assets were reported for, because
    a year in which only one obscure metric resolved is not a period the
    reader can compare anything across.
    """
    anchors = ("income.revenue", "balance.total_assets", "income.net_income")
    years = {
        year
        for (metric, year) in index.by_metric_year
        if metric in anchors
    }
    if not years:
        years = {year for (_, year) in index.by_metric_year}
    ordered = sorted(years, reverse=True)[:MAX_TABLE_PERIODS]
    return tuple(sorted(ordered))


def _currency(facts: Sequence[Fact]) -> str | None:
    """The currency the filer reports in, taken from the facts themselves."""
    for fact in facts:
        unit = fact.unit
        if unit and len(unit) == 3 and unit.isalpha() and unit.isupper():
            return unit
    return None


def _period_label(year: int) -> str:
    return f"FY{year}"


# --- Public: the browser document --------------------------------------------


def assemble_document(inputs: AssemblyInput) -> ReportDocument:
    """Builds the document every renderer works from.

    Facts appear in `facts` in first-appearance order, which is the order the
    provenance rail assigns its markers in — so marker 1 is the first source
    the reader meets rather than whichever fact happened to sort first.
    """
    index = _build_index(inputs.facts)
    cited: list[str] = []

    sections = tuple(
        section
        for section in (
            _snapshot_section(inputs, index, cited),
            _business_section(inputs, index, cited),
            _financials_section(inputs, index, cited),
            _analysis_section(inputs, index, cited),
            _peers_section(inputs, index, cited),
            _developments_section(inputs, cited),
            _risks_section(inputs, index, cited),
        )
    )

    ordered = _facts_in_citation_order(inputs.facts, cited)

    document = ReportDocument(
        report=inputs.report,
        company=inputs.company,
        depth=inputs.depth,
        facts=tuple(
            DocumentFact.of(fact, inputs.report.id) for fact in ordered
        ),
        filings=tuple(inputs.filings),
        sections=sections,
        charts=_charts(inputs, index),
        compliance=inputs.compliance,
        artifacts=tuple(inputs.artifacts),
        source_notes=tuple(inputs.source_notes),
    )

    logger.info(
        "Document assembled",
        extra={
            "report_id": inputs.report.id,
            "sections": len(sections),
            "cited_facts": len(document.facts),
            "periods": len(index.years),
        },
    )
    return document


def _facts_in_citation_order(
    facts: Sequence[Fact], cited: Sequence[str]
) -> list[Fact]:
    """Cited facts first in document order, then the rest.

    The remainder is kept rather than dropped: a fact that no table cites is
    still part of what was extracted and verified, and the workbook's sources
    tab lists all of it. Dropping it would make the compliance count and the
    document disagree.
    """
    by_id = {fact.fact_id: fact for fact in facts}
    ordered: list[Fact] = []
    seen: set[str] = set()

    for fact_id in cited:
        fact = by_id.get(fact_id)
        if fact is not None and fact_id not in seen:
            seen.add(fact_id)
            ordered.append(fact)

    for fact in facts:
        if fact.fact_id not in seen:
            seen.add(fact.fact_id)
            ordered.append(fact)

    return ordered


# --- Sections ----------------------------------------------------------------


def _prose_blocks(
    inputs: AssemblyInput, section_id: SectionId, cited: list[str]
) -> tuple[ProseBlock, ...]:
    """Generated prose for one section, with its citations recorded."""
    if inputs.prose is None:
        return ()
    written = inputs.prose.section(str(section_id))
    if written is None or not written.sentences:
        return ()

    blocks: list[ProseBlock] = []
    for position, sentence in enumerate(written.sentences, start=1):
        cited.extend(sentence.fact_ids)
        blocks.append(
            ProseBlock(
                id=f"{section_id}-{position}",
                text=sentence.text,
                fact_ids=tuple(sentence.fact_ids),
            )
        )
    return tuple(blocks)


def _table(
    table_id: str,
    caption: str,
    rows: Sequence[_Row],
    index: _Index,
    cited: list[str],
    *,
    unit_note: str,
) -> FigureTable | None:
    """One figure table, or None when not a single row has a figure."""
    periods = index.years
    if not periods:
        return None

    built: list[FigureRow] = []
    for row in rows:
        fact_ids: list[str | None] = []
        found = False
        for year in periods:
            fact = index.fact(row.metric, year)
            if fact is None:
                fact_ids.append(None)
                continue
            found = True
            fact_ids.append(fact.fact_id)
            cited.append(fact.fact_id)
        if found:
            built.append(
                FigureRow(
                    label=row.label,
                    fact_ids=tuple(fact_ids),
                    emphasis=row.emphasis,
                )
            )

    if not built:
        return None

    return FigureTable(
        id=table_id,
        caption=caption,
        periods=tuple(_period_label(year) for year in periods),
        rows=tuple(built),
        unit_note=unit_note,
    )


def _currency_note(index: _Index) -> str:
    return DISPLAY_SCALE_NOTE.format(currency=index.currency or "Reporting")


def _snapshot_section(
    inputs: AssemblyInput, index: _Index, cited: list[str]
) -> DocumentSection:
    """Identity and market pricing. Tier 3 lives here and nowhere earlier."""
    prose = _prose_blocks(inputs, SectionId.SNAPSHOT, cited)

    rows: list[FigureRow] = []
    for row in _MARKET_ROWS:
        fact = index.latest.get(row.metric)
        if fact is None:
            continue
        cited.append(fact.fact_id)
        rows.append(FigureRow(label=row.label, fact_ids=(fact.fact_id,)))

    tables: tuple[FigureTable, ...] = ()
    if rows:
        as_of = index.latest["market.price"].period_end.isoformat()
        tables = (
            FigureTable(
                id="snapshot-market",
                caption="Market data",
                periods=("Current",),
                rows=tuple(rows),
                unit_note=(
                    f"Tier 3, from a market data provider, as at {as_of}. "
                    "Never used to support a figure in the financial "
                    "highlights."
                ),
            ),
        )

    return DocumentSection(
        id=SectionId.SNAPSHOT,
        title=_SECTION_TITLES[SectionId.SNAPSHOT],
        unavailable_reason=(
            None if (prose or tables) else _UNAVAILABLE[SectionId.SNAPSHOT]
        ),
        prose=prose,
        tables=tables,
    )


def _business_section(
    inputs: AssemblyInput, index: _Index, cited: list[str]
) -> DocumentSection:
    """The filer's own description of what it sells."""
    prose = list(_prose_blocks(inputs, SectionId.BUSINESS, cited))

    # Where no prose was generated, the filer's own words are quoted directly
    # rather than leaving the section empty. This is the filing, not a
    # summary — and Item 1 and Item 7 are two different filing items, so each
    # quote carries its own label rather than reading as one undivided block.
    if not prose:
        labels = {
            "business.description": "From item 1. Business",
            "business.mdna": "From item 7. Management's discussion and analysis",
        }
        for metric, label in labels.items():
            fact = index.latest.get(metric)
            if fact is None:
                continue
            cited.append(fact.fact_id)
            prose.append(
                ProseBlock(
                    id=f"business-{metric.rsplit('.', 1)[-1]}",
                    text=m04_narrative.summary_of(fact),
                    fact_ids=(fact.fact_id,),
                    label=label,
                )
            )

    return DocumentSection(
        id=SectionId.BUSINESS,
        title=_SECTION_TITLES[SectionId.BUSINESS],
        unavailable_reason=(
            None if prose else _UNAVAILABLE[SectionId.BUSINESS]
        ),
        prose=tuple(prose),
    )


def _financials_section(
    inputs: AssemblyInput, index: _Index, cited: list[str]
) -> DocumentSection:
    """Reported figures only. Tier 1 and 2, enforced upstream by m06."""
    note = _currency_note(index)
    tables = [
        table
        for table in (
            _table(
                "financials-income",
                "Income statement",
                _INCOME_ROWS,
                index,
                cited,
                unit_note=note,
            ),
            _table(
                "financials-balance",
                "Balance sheet",
                _BALANCE_ROWS,
                index,
                cited,
                unit_note=note,
            ),
            _table(
                "financials-cashflow",
                "Cash flow",
                _CASHFLOW_ROWS,
                index,
                cited,
                unit_note=note,
            ),
        )
        if table is not None
    ]

    return DocumentSection(
        id=SectionId.FINANCIALS,
        title=_SECTION_TITLES[SectionId.FINANCIALS],
        unavailable_reason=(
            None if tables else _UNAVAILABLE[SectionId.FINANCIALS]
        ),
        prose=_prose_blocks(inputs, SectionId.FINANCIALS, cited),
        tables=tuple(tables),
    )


def _analysis_section(
    inputs: AssemblyInput, index: _Index, cited: list[str]
) -> DocumentSection:
    """Everything computed. Each row carries its formula on the fact."""
    derived_note = (
        "Every figure below is calculated. The formula travels with each one."
    )

    tables = [
        table
        for table in (
            _table(
                "analysis-margins",
                "Margins and returns",
                _MARGIN_ROWS,
                index,
                cited,
                unit_note=derived_note,
            ),
            _table(
                "analysis-strength",
                "Liquidity and leverage",
                _STRENGTH_ROWS,
                index,
                cited,
                unit_note=derived_note,
            ),
            _table(
                "analysis-growth",
                "Growth",
                _GROWTH_ROWS,
                index,
                cited,
                unit_note=derived_note,
            ),
        )
        if table is not None
    ]

    for prefix, caption, rows in _SECTOR_ROWS:
        if not any(fact.startswith(f"{prefix}.") for fact in index.latest):
            continue
        sector_table = _table(
            f"analysis-{prefix}",
            caption,
            rows,
            index,
            cited,
            unit_note=derived_note,
        )
        if sector_table is not None:
            tables.append(sector_table)

    segments = _segment_table(index, cited)
    if segments is not None:
        tables.append(segments)

    return DocumentSection(
        id=SectionId.ANALYSIS,
        title=_SECTION_TITLES[SectionId.ANALYSIS],
        unavailable_reason=(
            None if tables else _UNAVAILABLE[SectionId.ANALYSIS]
        ),
        prose=_prose_blocks(inputs, SectionId.ANALYSIS, cited),
        tables=tuple(tables),
    )


def _segment_series(facts: Sequence[Fact]) -> dict[str, dict[int, Fact]]:
    """Segment revenue by member label and fiscal year, in disclosure order."""
    grouped: dict[str, dict[int, Fact]] = {}
    for fact in facts:
        if fact.metric != SEGMENT_METRIC or fact.segment_member is None:
            continue
        if fact.fiscal_year is None:
            continue
        label = fact.segment_label or fact.segment_member
        grouped.setdefault(label, {})[fact.fiscal_year] = fact
    return grouped


def _segment_table(index: _Index, cited: list[str]) -> FigureTable | None:
    """Revenue by segment, as the filer reported it."""
    grouped = index.segments
    if not grouped or not index.years:
        return None

    rows: list[FigureRow] = []
    for label, by_year in grouped.items():
        fact_ids: list[str | None] = []
        found = False
        for year in index.years:
            fact = by_year.get(year)
            if fact is None:
                fact_ids.append(None)
                continue
            found = True
            fact_ids.append(fact.fact_id)
            cited.append(fact.fact_id)
        if found:
            rows.append(FigureRow(label=label, fact_ids=tuple(fact_ids)))

    if not rows:
        return None

    return FigureTable(
        id="analysis-segments",
        caption="Revenue by segment",
        periods=tuple(_period_label(year) for year in index.years),
        rows=tuple(rows),
        unit_note=(
            "As the filer reported it. Segments sum to consolidated revenue "
            "within the tolerance stated in the compliance strip."
        ),
    )


def _peers_section(
    inputs: AssemblyInput, index: _Index, cited: list[str]
) -> DocumentSection:
    """Comparable companies: how each was chosen, and how they compare."""
    rows: list[FigureRow] = []
    for metric, fact in sorted(index.latest.items()):
        # Scoped to the selection-note facts specifically. The comparison
        # table's own facts also start with "peer." and are rendered
        # separately by `_peer_comparison_table`, not as a flat list here.
        if not metric.startswith("peer.company."):
            continue
        cited.append(fact.fact_id)
        rows.append(
            FigureRow(label=fact.label, fact_ids=(fact.fact_id,))
        )

    tables: list[FigureTable] = []
    if rows:
        note = "How each peer was selected is stated beside it."
        if inputs.peers is not None and inputs.peers.has_fiscal_year_mismatch:
            note = (
                f"{note} One or more peers closes its books at a materially "
                "different time, which is disclosed on the peer."
            )
        tables.append(
            FigureTable(
                id="peers-list",
                caption="Comparable companies",
                periods=("Selection",),
                rows=tuple(rows),
                unit_note=note,
            )
        )

    comparison_table = _peer_comparison_table(inputs.peer_comparison, index, cited)
    if comparison_table is not None:
        tables.append(comparison_table)

    return DocumentSection(
        id=SectionId.PEERS,
        title=_SECTION_TITLES[SectionId.PEERS],
        unavailable_reason=None if tables else _UNAVAILABLE[SectionId.PEERS],
        prose=_prose_blocks(inputs, SectionId.PEERS, cited),
        tables=tuple(tables),
    )


def _peer_comparison_table(
    comparison: m08_peers.PeerComparison | None, index: _Index, cited: list[str]
) -> FigureTable | None:
    """Revenue, margin and growth for the subject and each peer, side by side.

    Columns are companies rather than fiscal periods — `FigureTable.periods`
    is reused as the column headers, the same way `peers-list` above reuses it
    for "Selection". Each row's `fact_ids` line up positionally with those
    columns; a company missing a figure gets None in that slot; the frontend
    already reads a None fact id as a blank cell, not a zero.
    """
    if comparison is None or not comparison.rows:
        return None

    metric_rows: tuple[
        tuple[str, Callable[[m08_peers.PeerComparisonRow], Fact | None]], ...
    ] = (
        ("Revenue", lambda row: row.revenue),
        ("Net margin", lambda row: row.net_margin),
        ("Operating margin", lambda row: row.operating_margin),
        ("Revenue growth, year on year", lambda row: row.revenue_growth),
    )

    rows: list[FigureRow] = []
    for label, getter in metric_rows:
        fact_ids: list[str | None] = []
        any_present = False
        for row in comparison.rows:
            fact = getter(row)
            if fact is None:
                fact_ids.append(None)
                continue
            cited.append(fact.fact_id)
            fact_ids.append(fact.fact_id)
            any_present = True
        if any_present:
            rows.append(FigureRow(label=label, fact_ids=tuple(fact_ids)))

    if not rows:
        return None

    note = (
        f"{_currency_note(index)}. Each company's own figures, read from its "
        "own SEC filings the same way the subject's are."
    )
    if comparison.notes:
        note = f"{note} {' '.join(comparison.notes)}"

    return FigureTable(
        id="peers-comparison",
        caption="Financial comparison",
        periods=tuple(row.ticker for row in comparison.rows),
        rows=tuple(rows),
        unit_note=note,
    )


def _developments_section(
    inputs: AssemblyInput, cited: list[str]
) -> DocumentSection:
    """The filed current reports, most recent first."""
    for event in inputs.events:
        cited.extend(event.fact_ids)

    return DocumentSection(
        id=SectionId.DEVELOPMENTS,
        title=_SECTION_TITLES[SectionId.DEVELOPMENTS],
        unavailable_reason=(
            None if inputs.events else _UNAVAILABLE[SectionId.DEVELOPMENTS]
        ),
        prose=_prose_blocks(inputs, SectionId.DEVELOPMENTS, cited),
        events=tuple(inputs.events),
    )


def _risks_section(
    inputs: AssemblyInput, index: _Index, cited: list[str]
) -> DocumentSection:
    """The filer's own risk headings, in the order the filer wrote them."""
    risk_fact = index.latest.get(m04_narrative.RISK_METRIC)
    risks: tuple[RiskItem, ...] = ()

    if risk_fact is not None:
        cited.append(risk_fact.fact_id)
        headings = m04_narrative.risk_headings([risk_fact])
        risks = tuple(
            RiskItem(
                id=f"risk-{position}",
                heading=heading,
                summary=(
                    "Disclosed by the filer in the risk factors item of its "
                    "annual report."
                ),
                fact_ids=(risk_fact.fact_id,),
            )
            for position, heading in enumerate(headings, start=1)
        )

    prose = _prose_blocks(inputs, SectionId.RISKS, cited)

    return DocumentSection(
        id=SectionId.RISKS,
        title=_SECTION_TITLES[SectionId.RISKS],
        unavailable_reason=(
            None if (risks or prose) else _UNAVAILABLE[SectionId.RISKS]
        ),
        prose=prose,
        risks=risks,
    )


# --- Charts ------------------------------------------------------------------
# Values arrive in their display scale, computed here. The browser plots them;
# it never divides, sums or derives a figure.


def _scaled(fact: Fact | None) -> float | None:
    """A fact's value in the scale the charts and tables are drawn in."""
    if fact is None or fact.value is None:
        return None
    if any(fact.metric.startswith(p) for p in UNSCALED_METRIC_PREFIXES):
        return round(fact.value, 4)
    return round(fact.value / DISPLAY_SCALE_DIVISOR, 3)


def _charts(inputs: AssemblyInput, index: _Index) -> ReportCharts:
    """Every series the document can support. A series with no data is None."""
    return ReportCharts(
        revenue_margin=_revenue_margin_series(index),
        segment_mix=_segment_mix_series(index),
        cash_flow=_cash_flow_series(index),
        peer_valuation=_peer_valuation_series(inputs.peer_comparison),
    )


def _peer_valuation_series(
    comparison: m08_peers.PeerComparison | None,
) -> PeerValuationSeries | None:
    """Price to earnings and EV to EBITDA for the subject and its peers.

    Values are already the multiple m08 computed — "times", not currency —
    so they are read straight off the fact rather than passed through
    `_scaled`, which would apply the currency-millions scaling this series
    does not use.
    """
    if comparison is None:
        return None

    points: list[PeerValuationPoint] = []
    ids: list[str] = []

    for row in comparison.rows:
        pe, ev = row.price_to_earnings, row.ev_to_ebitda
        if pe is None and ev is None:
            continue
        ids.extend(fact.fact_id for fact in (pe, ev) if fact is not None)
        points.append(
            PeerValuationPoint(
                ticker=row.ticker,
                name=row.name,
                is_subject=row.is_subject,
                price_to_earnings=pe.value if pe is not None else None,
                ev_to_ebitda=ev.value if ev is not None else None,
            )
        )

    if not points:
        return None

    return PeerValuationSeries(
        meta=SeriesMeta(unit_label="Times", fact_ids=tuple(ids)),
        points=tuple(points),
    )


def _revenue_margin_series(index: _Index) -> RevenueMarginSeries | None:
    if not index.years:
        return None

    points: list[RevenueMarginPoint] = []
    ids: list[str] = []
    populated = False

    for year in index.years:
        revenue = index.fact("income.revenue", year)
        gross = index.fact("margin.gross", year)
        operating = index.fact("margin.operating", year)
        ids.extend(
            fact.fact_id for fact in (revenue, gross, operating) if fact
        )
        point = RevenueMarginPoint(
            period=_period_label(year),
            revenue=_scaled(revenue),
            gross_margin_pct=_scaled(gross),
            operating_margin_pct=_scaled(operating),
        )
        populated = populated or point.revenue is not None
        points.append(point)

    if not populated:
        return None

    return RevenueMarginSeries(
        meta=SeriesMeta(
            unit_label=f"{index.currency or 'Reporting'} millions",
            fact_ids=tuple(ids),
        ),
        points=tuple(points),
    )


def _segment_mix_series(index: _Index) -> SegmentMixSeries | None:
    grouped = index.segments
    if not grouped or not index.years:
        return None

    # Ordered by the most recent year's size so the largest band is drawn
    # first and the colour order is stable across periods.
    newest = index.years[-1]

    def newest_value(entry: tuple[str, dict[int, Fact]]) -> float:
        fact = entry[1].get(newest)
        return -(fact.value or 0.0) if fact is not None else 0.0

    ranked = sorted(grouped.items(), key=newest_value)[:MAX_CHART_SEGMENTS]

    ids: list[str] = []
    points: list[SegmentMixPoint] = []
    for year in index.years:
        values: dict[str, float] = {}
        for label, by_year in ranked:
            scaled = _scaled(by_year.get(year))
            if scaled is None:
                continue
            values[label] = scaled
            fact = by_year.get(year)
            if fact is not None:
                ids.append(fact.fact_id)
        points.append(
            SegmentMixPoint(period=_period_label(year), values=values)
        )

    if not any(point.values for point in points):
        return None

    return SegmentMixSeries(
        meta=SeriesMeta(
            unit_label=f"{index.currency or 'Reporting'} millions",
            fact_ids=tuple(ids),
        ),
        segments=tuple(label for label, _ in ranked),
        points=tuple(points),
    )


def _cash_flow_series(index: _Index) -> CashFlowSeries | None:
    if not index.years:
        return None

    ids: list[str] = []
    points: list[CashFlowPoint] = []
    populated = False

    for year in index.years:
        operating = index.fact("cashflow.operating", year)
        capex = index.fact("cashflow.capital_expenditure", year)
        free = index.fact("cashflow.free_cash_flow", year)
        ids.extend(fact.fact_id for fact in (operating, capex, free) if fact)
        point = CashFlowPoint(
            period=_period_label(year),
            operating_cash_flow=_scaled(operating),
            capex=_scaled(capex),
            free_cash_flow=_scaled(free),
        )
        populated = populated or point.operating_cash_flow is not None
        points.append(point)

    if not populated:
        return None

    return CashFlowSeries(
        meta=SeriesMeta(
            unit_label=f"{index.currency or 'Reporting'} millions",
            fact_ids=tuple(ids),
        ),
        points=tuple(points),
    )


# --- PDF ---------------------------------------------------------------------
# The printed trisheet. Paper-white, hairline rules, figures right-aligned in
# a monospace face — a research document, not a dashboard. Every figure carries
# a superscript marker resolved in the sources appendix, which is the printed
# form of the provenance rail.

#: Typography. The named faces are the design system's; the generics after each
#: are what a machine without them falls back to. Rendering happens with no
#: network access, so a face that is not installed is simply not used — the
#: document still sets correctly.
_PDF_STYLESHEET = """
@page {
  size: A4;
  margin: 18mm 16mm 20mm 16mm;
  @bottom-left {
    content: "Trisheet — sourced from SEC filings";
    font-family: "IBM Plex Sans", "Segoe UI", sans-serif;
    font-size: 7pt;
    color: #6c7671;
  }
  @bottom-right {
    content: counter(page) " / " counter(pages);
    font-family: "IBM Plex Mono", "Consolas", monospace;
    font-size: 7pt;
    color: #6c7671;
  }
}

:root {
  --paper: #FBFAF7;
  --ink: #14201C;
  --rule: #E2DED4;
  --certified: #1F4D3D;
  --market: #3E5C7A;
  --flag: #9E3B26;
  --muted: #6c7671;
}

body {
  background: #FBFAF7;
  color: #14201C;
  font-family: "IBM Plex Sans", "Segoe UI", sans-serif;
  font-size: 8.5pt;
  line-height: 1.5;
}

h1, h2, h3 {
  font-family: "Fraunces", Georgia, "Times New Roman", serif;
  font-weight: 600;
  color: #14201C;
  margin: 0;
}

h1 { font-size: 22pt; letter-spacing: -0.01em; }
h2 { font-size: 12pt; margin-top: 16pt; }
h3 { font-size: 9pt; margin-top: 10pt; }

.masthead { border-bottom: 1.5px solid #14201C; padding-bottom: 8pt; }
.masthead .ticker {
  font-family: "IBM Plex Mono", "Consolas", monospace;
  font-size: 10pt;
  color: #1F4D3D;
  letter-spacing: 0.06em;
}
.masthead .meta {
  font-family: "IBM Plex Mono", "Consolas", monospace;
  font-size: 7.5pt;
  color: #6c7671;
  margin-top: 4pt;
}

.compliance {
  border-bottom: 1px solid #E2DED4;
  padding: 6pt 0;
  font-size: 7.5pt;
  color: #6c7671;
}
.compliance .figure {
  font-family: "IBM Plex Mono", "Consolas", monospace;
  color: #14201C;
}
.compliance .passed { color: #1F4D3D; }
.compliance .failed { color: #9E3B26; }

section { break-inside: auto; margin-top: 4pt; }
section > h2 {
  border-bottom: 1px solid #14201C;
  padding-bottom: 3pt;
}

p { margin: 6pt 0; }
p.prose { text-align: justify; }

.unavailable {
  color: #6c7671;
  font-style: italic;
}

table {
  width: 100%;
  border-collapse: collapse;
  margin-top: 8pt;
  break-inside: avoid;
}
caption {
  caption-side: top;
  text-align: left;
  font-family: "Fraunces", Georgia, serif;
  font-size: 9pt;
  padding-bottom: 2pt;
}
th, td {
  border-bottom: 0.5px solid #E2DED4;
  padding: 2.5pt 4pt;
  vertical-align: baseline;
}
thead th {
  border-bottom: 1px solid #14201C;
  font-family: "IBM Plex Mono", "Consolas", monospace;
  font-size: 7pt;
  font-weight: 500;
  text-transform: uppercase;
  letter-spacing: 0.05em;
  text-align: right;
  color: #6c7671;
}
thead th:first-child { text-align: left; }
td.figure {
  font-family: "IBM Plex Mono", "Consolas", monospace;
  text-align: right;
  font-variant-numeric: tabular-nums;
  white-space: nowrap;
}
td.label { width: 34%; }
tr.total td { font-weight: 600; }
tr.derived td.label { color: #1F4D3D; }
td.absent { color: #6c7671; }

.unit-note {
  font-size: 7pt;
  color: #6c7671;
  margin: 2pt 0 0 0;
}

sup.marker {
  font-family: "IBM Plex Mono", "Consolas", monospace;
  font-size: 6pt;
  color: #1F4D3D;
  padding-left: 1pt;
}

ol.risks { margin: 6pt 0 0 0; padding-left: 14pt; }
ol.risks li { margin-bottom: 4pt; }

ul.events { list-style: none; margin: 6pt 0 0 0; padding: 0; }
ul.events li {
  border-bottom: 0.5px solid #E2DED4;
  padding: 4pt 0;
}
ul.events .when {
  font-family: "IBM Plex Mono", "Consolas", monospace;
  font-size: 7.5pt;
  color: #6c7671;
}

table.sources { font-size: 7pt; }
table.sources td, table.sources th { padding: 2pt 3pt; }
table.sources .mono {
  font-family: "IBM Plex Mono", "Consolas", monospace;
  word-break: break-all;
}
.tier-1 { color: #1F4D3D; }
.tier-2 { color: #1F4D3D; }
.tier-3 { color: #3E5C7A; }
.tier-4 { color: #9E3B26; }
"""

#: Tier labels for the printed sources appendix.
_TIER_LABELS: dict[int, str] = {
    1: "Tier 1 · filing",
    2: "Tier 2 · company",
    3: "Tier 3 · market data",
    4: "Tier 4 · news",
}


@dataclass(frozen=True, slots=True)
class _SourceCard:
    """One entry in the printed sources appendix."""

    marker: int
    accession_no: str
    tier: int
    form: str | None
    filed_date: dt.date
    url: str
    fact_ids: tuple[str, ...]


def _source_cards(document: ReportDocument) -> list[_SourceCard]:
    """Sources in first-appearance order, one card per accession number.

    The same grouping the browser's provenance rail uses: two figures drawn
    from the same filing share one marker, because they share one source.
    """
    forms = {
        filing.accession_no: filing.form for filing in document.filings
    }
    order: list[str] = []
    grouped: dict[str, list[DocumentFact]] = {}

    for fact in document.facts:
        grouped.setdefault(fact.accession_no, [])
        if fact.accession_no not in order:
            order.append(fact.accession_no)
        grouped[fact.accession_no].append(fact)

    cards: list[_SourceCard] = []
    for marker, accession in enumerate(order, start=1):
        members = grouped[accession]
        first = members[0]
        cards.append(
            _SourceCard(
                marker=marker,
                accession_no=accession,
                tier=int(first.tier),
                form=forms.get(accession),
                filed_date=first.filed_date,
                url=str(first.source_url),
                fact_ids=tuple(fact.id for fact in members),
            )
        )
    return cards


def _marker_index(cards: Sequence[_SourceCard]) -> dict[str, int]:
    """Fact id to the superscript marker that resolves it."""
    return {
        fact_id: card.marker for card in cards for fact_id in card.fact_ids
    }


def _escape(text: str) -> str:
    return html.escape(text, quote=True)


def render_html(document: ReportDocument) -> str:
    """The printable document, as a self-contained HTML string.

    Separated from `assemble_pdf` because it is pure and therefore testable:
    the layout can be asserted on without a rendering engine installed, and
    PDF generation becomes a single call over a value this function returns.
    """
    facts = {fact.id: fact for fact in document.facts}
    cards = _source_cards(document)
    markers = _marker_index(cards)

    parts: list[str] = [
        f"<style>{_PDF_STYLESHEET}</style>",
        _render_masthead(document),
        _render_compliance(document),
    ]

    for section in document.sections:
        parts.append(_render_section(section, facts, markers))

    parts.append(_render_sources(cards))

    title = _escape(f"{document.company.name} — Trisheet")
    return (
        "<!DOCTYPE html><html lang='en'><head><meta charset='utf-8'>"
        f"<title>{title}</title></head><body>"
        + "".join(parts)
        + "</body></html>"
    )


def _render_masthead(document: ReportDocument) -> str:
    company = document.company
    completed = document.report.completed_at or document.report.created_at
    meta = " · ".join(
        part
        for part in (
            f"CIK {company.cik}",
            company.sector,
            _filer_form(company),
            company.reporting_currency,
            f"Generated {completed.date().isoformat()}",
        )
        if part
    )
    return (
        "<header class='masthead'>"
        f"<div class='ticker'>{_escape(company.ticker)}</div>"
        f"<h1>{_escape(company.name)}</h1>"
        f"<div class='meta'>{_escape(meta)}</div>"
        "</header>"
    )


def _filer_form(company: Company) -> str:
    return {
        "domestic": "Files 10-K",
        "foreign": "Files 20-F",
        "canadian": "Files 40-F",
    }[str(company.filer_type)]


def _render_compliance(document: ReportDocument) -> str:
    compliance = document.compliance
    state = "passed" if compliance.passed else "failed"
    verdict = (
        "Every figure traced to a filing"
        if compliance.passed
        else "Verification found blocking issues"
    )
    tiers = " ".join(
        f"T{tier} <span class='figure'>{count}</span>"
        for tier, count in sorted(compliance.tier_counts.items())
    )
    return (
        "<div class='compliance'>"
        f"<span class='{state}'>{_escape(verdict)}</span> · "
        f"Citation coverage <span class='figure'>"
        f"{_escape(compliance.coverage_display)}</span> "
        f"({compliance.cited_figure_count}/{compliance.figure_count} figures)"
        f" · Facts <span class='figure'>{compliance.fact_count}</span>"
        f" · {tiers}"
        f" · Verified {compliance.verified_at.date().isoformat()}"
        "</div>"
    )


def _render_section(
    section: DocumentSection,
    facts: Mapping[str, DocumentFact],
    markers: Mapping[str, int],
) -> str:
    body: list[str] = [f"<h2>{_escape(section.title)}</h2>"]

    if section.unavailable_reason:
        body.append(
            f"<p class='unavailable'>{_escape(section.unavailable_reason)}</p>"
        )

    for block in section.prose:
        body.append(
            "<p class='prose'>"
            + _escape(block.text)
            + _markers_for(block.fact_ids, markers)
            + "</p>"
        )

    for table in section.tables:
        body.append(_render_table(table, facts, markers))

    if section.events:
        body.append(_render_events(section.events, markers))

    if section.risks:
        body.append(_render_risks(section.risks, markers))

    return f"<section id='{_escape(str(section.id))}'>" + "".join(body) + "</section>"


def _markers_for(
    fact_ids: Iterable[str], markers: Mapping[str, int]
) -> str:
    """Superscripts for a set of facts, deduplicated and in order."""
    numbers = sorted(
        {markers[fact_id] for fact_id in fact_ids if fact_id in markers}
    )
    if not numbers:
        return ""
    return "<sup class='marker'>" + ",".join(str(n) for n in numbers) + "</sup>"


def _render_table(
    table: FigureTable,
    facts: Mapping[str, DocumentFact],
    markers: Mapping[str, int],
) -> str:
    header = "".join(f"<th>{_escape(period)}</th>" for period in table.periods)
    rows: list[str] = []

    for row in table.rows:
        cells: list[str] = []
        for fact_id in row.fact_ids:
            fact = facts.get(fact_id) if fact_id else None
            if fact is None:
                cells.append(
                    f"<td class='figure absent'>{_escape(NOT_DISCLOSED_TEXT)}</td>"
                )
                continue
            marker = _markers_for((fact.id,), markers)
            cells.append(
                "<td class='figure'>"
                + _escape(fact.display_value)
                + marker
                + "</td>"
            )
        classes = f" class='{row.emphasis}'" if row.emphasis else ""
        rows.append(
            f"<tr{classes}><td class='label'>{_escape(row.label)}</td>"
            + "".join(cells)
            + "</tr>"
        )

    return (
        "<table>"
        f"<caption>{_escape(table.caption)}</caption>"
        f"<thead><tr><th>{_escape('Figure')}</th>{header}</tr></thead>"
        f"<tbody>{''.join(rows)}</tbody>"
        "</table>"
        f"<p class='unit-note'>{_escape(table.unit_note)}</p>"
    )


def _render_events(
    events: Sequence[DevelopmentEvent], markers: Mapping[str, int]
) -> str:
    items = "".join(
        "<li>"
        f"<span class='when'>{event.date.isoformat()} · "
        f"{_escape(event.form)}"
        + (f" · {_escape(', '.join(event.items))}" if event.items else "")
        + "</span><br>"
        + _escape(event.headline)
        + _markers_for(event.fact_ids, markers)
        + "</li>"
        for event in events
    )
    return f"<ul class='events'>{items}</ul>"


def _render_risks(
    risks: Sequence[RiskItem], markers: Mapping[str, int]
) -> str:
    items = "".join(
        "<li>"
        + _escape(risk.heading)
        + _markers_for(risk.fact_ids, markers)
        + "</li>"
        for risk in risks
    )
    return f"<ol class='risks'>{items}</ol>"


def _render_sources(cards: Sequence[_SourceCard]) -> str:
    """The printed provenance rail.

    It is an appendix on paper because a printed page has no hover, but it is
    the same content and the same markers — the sources travel with the
    document rather than being a feature of the screen it was read on.
    """
    if not cards:
        return ""

    rows = "".join(
        "<tr>"
        f"<td class='mono'>{card.marker}</td>"
        f"<td class='tier-{card.tier}'>{_escape(_TIER_LABELS[card.tier])}</td>"
        f"<td>{_escape(card.form or 'Not a filing')}</td>"
        f"<td class='mono'>{_escape(card.accession_no)}</td>"
        f"<td class='mono'>{card.filed_date.isoformat()}</td>"
        f"<td class='mono'>{_escape(card.url)}</td>"
        "</tr>"
        for card in cards
    )
    return (
        "<section id='sources'><h2>Sources</h2>"
        "<table class='sources'><thead><tr>"
        "<th>#</th><th>Tier</th><th>Form</th><th>Accession</th>"
        "<th>Filed</th><th>Document</th>"
        "</tr></thead>"
        f"<tbody>{rows}</tbody></table>"
        "</section>"
    )


def assemble_pdf(document: ReportDocument) -> bytes:
    """Renders the report as a PDF.

    Raises:
        ArtifactRenderError: WeasyPrint is not installed, or its native text
            and layout libraries are unavailable on this machine. The message
            is written for a reader, because it is what the interface shows in
            place of a download link.
    """
    markup = render_html(document)

    try:
        from weasyprint import HTML
    except Exception as cause:  # noqa: BLE001 — an import failure is a finding
        # WeasyPrint raises OSError, not ImportError, when its pango and
        # cairo libraries are missing — so the whole import is guarded.
        logger.warning(
            "WeasyPrint is unavailable; the PDF was not rendered",
            extra={"error": str(cause)},
        )
        message = (
            "The PDF renderer is not available on this server, so this "
            "report was not published as a PDF. The figures are unaffected."
        )
        raise ArtifactRenderError(message) from cause

    try:
        rendered = HTML(string=markup).write_pdf()
    except Exception as cause:  # noqa: BLE001 — surfaced as a typed failure
        logger.error("PDF rendering failed", exc_info=cause)
        message = "This report could not be rendered as a PDF."
        raise ArtifactRenderError(message) from cause

    if not rendered:
        message = "The PDF renderer produced an empty document."
        raise ArtifactRenderError(message)

    return bytes(rendered)


# --- XLSX --------------------------------------------------------------------
# The workbook is the auditable form of the report. The statement tabs hold the
# reported figures and nothing else; the analysis tab holds no values at all,
# only formulas pointing at those cells. Change a reported figure and every
# ratio moves — which is the point: a reader who disagrees with a margin can
# see the division rather than take it on trust.

_SHEET_COVER = "Trisheet"
_SHEET_INCOME = "Income statement"
_SHEET_BALANCE = "Balance sheet"
_SHEET_CASHFLOW = "Cash flow"
_SHEET_ANALYSIS = "Analysis"
_SHEET_SEGMENTS = "Segments"
_SHEET_ASSUMPTIONS = "Assumptions"
_SHEET_SOURCES = "Sources"

#: Where the days-in-year assumption lives, so working-capital formulas can
#: reference it rather than hard-coding 365.
_ASSUMPTION_DAYS_CELL = f"'{_SHEET_ASSUMPTIONS}'!$B$4"

_FORMAT_CURRENCY = "#,##0;(#,##0)"
_FORMAT_PERCENT = '0.0"%"'
_FORMAT_MULTIPLE = '0.00"x"'
_FORMAT_DAYS = '0" days"'
_FORMAT_PER_SHARE = "#,##0.00;(#,##0.00)"

#: Metrics whose reported value is already per-share or a count rather than a
#: currency amount, so they take a different number format.
_PER_SHARE_METRICS = frozenset(
    {
        "income.eps_basic",
        "income.eps_diluted",
        "income.dividends_per_share",
        "market.price",
        "market.previous_close",
        "market.fifty_two_week_high",
        "market.fifty_two_week_low",
    }
)


@dataclass(frozen=True, slots=True)
class _Derived:
    """One analysis row: a label, an Excel formula and how to format it.

    `formula` is a template taking `{cell}` lookups resolved per period. It is
    written once here and evaluated for every column, so a ratio cannot be
    defined differently in two years of the same table.
    """

    label: str
    #: Called with a resolver that maps "metric" or "metric@-1" to a cell
    #: reference, or None when that figure was not reported.
    build: _FormulaBuilder
    number_format: str
    note: str


def _formula_rows() -> tuple[_Derived, ...]:
    """Every derived row in the workbook, with the formula behind it.

    Each builder returns None when an input it needs is missing for that
    period, which renders as "Not disclosed" rather than as a zero.
    """

    def ratio(
        numerator: str, denominator: str, scale: str = ""
    ) -> _FormulaBuilder:
        def build(at: _CellRef) -> str | None:
            top, bottom = at(numerator), at(denominator)
            if top is None or bottom is None:
                return None
            return f"=IFERROR({top}/{bottom}{scale},\"\")"

        return build

    def average_return(numerator: str, balance: str) -> _FormulaBuilder:
        """Return on an average balance, matching how m07 computes it."""

        def build(at: _CellRef) -> str | None:
            top, closing, opening = at(numerator), at(balance), at(f"{balance}@-1")
            if top is None or closing is None:
                return None
            base = closing if opening is None else f"AVERAGE({opening},{closing})"
            return f"=IFERROR({top}/({base})*100,\"\")"

        return build

    def growth(metric: str) -> _FormulaBuilder:
        def build(at: _CellRef) -> str | None:
            current, prior = at(metric), at(f"{metric}@-1")
            if current is None or prior is None:
                return None
            return f"=IFERROR(({current}/{prior}-1)*100,\"\")"

        return build

    def free_cash_flow(at: _CellRef) -> str | None:
        operating, capex = at("cashflow.operating"), at(
            "cashflow.capital_expenditure"
        )
        if operating is None or capex is None:
            return None
        # Capital expenditure is reported as a positive outflow, so it is
        # subtracted with ABS rather than added — a filer that signs it the
        # other way would otherwise double the cash flow.
        return f"=IFERROR({operating}-ABS({capex}),\"\")"

    def fcf_conversion(at: _CellRef) -> str | None:
        operating, capex, net = (
            at("cashflow.operating"),
            at("cashflow.capital_expenditure"),
            at("income.net_income"),
        )
        if operating is None or capex is None or net is None:
            return None
        return f"=IFERROR(({operating}-ABS({capex}))/{net}*100,\"\")"

    def quick_ratio(at: _CellRef) -> str | None:
        current, inventory, liabilities = (
            at("balance.current_assets"),
            at("balance.inventory"),
            at("balance.current_liabilities"),
        )
        if current is None or liabilities is None:
            return None
        stock = f"-{inventory}" if inventory else ""
        return f"=IFERROR(({current}{stock})/{liabilities},\"\")"

    def total_debt(at: _CellRef) -> str | None:
        short, long = at("balance.short_term_debt"), at("balance.long_term_debt")
        parts = [cell for cell in (short, long) if cell]
        if not parts:
            return None
        return "+".join(parts)

    def debt_to_equity(at: _CellRef) -> str | None:
        debt, equity = total_debt(at), at("balance.total_equity")
        if debt is None or equity is None:
            return None
        return f"=IFERROR(({debt})/{equity},\"\")"

    def days_sales_outstanding(at: _CellRef) -> str | None:
        receivables, revenue = (
            at("balance.accounts_receivable"),
            at("income.revenue"),
        )
        if receivables is None or revenue is None:
            return None
        return (
            f"=IFERROR({receivables}/{revenue}*{_ASSUMPTION_DAYS_CELL},\"\")"
        )

    def days_inventory(at: _CellRef) -> str | None:
        inventory, cost = (
            at("balance.inventory"),
            at("income.cost_of_revenue"),
        )
        if inventory is None or cost is None:
            return None
        return f"=IFERROR({inventory}/{cost}*{_ASSUMPTION_DAYS_CELL},\"\")"

    return (
        _Derived(
            "Gross margin",
            ratio("income.gross_profit", "income.revenue", "*100"),
            _FORMAT_PERCENT,
            "Gross profit ÷ revenue",
        ),
        _Derived(
            "Operating margin",
            ratio("income.operating_income", "income.revenue", "*100"),
            _FORMAT_PERCENT,
            "Operating income ÷ revenue",
        ),
        _Derived(
            "Net margin",
            ratio("income.net_income", "income.revenue", "*100"),
            _FORMAT_PERCENT,
            "Net income ÷ revenue",
        ),
        _Derived(
            "Effective tax rate",
            ratio(
                "income.income_tax_expense", "income.pretax_income", "*100"
            ),
            _FORMAT_PERCENT,
            "Income tax expense ÷ income before tax",
        ),
        _Derived(
            "Return on equity",
            average_return("income.net_income", "balance.total_equity"),
            _FORMAT_PERCENT,
            "Net income ÷ average total equity",
        ),
        _Derived(
            "Return on assets",
            average_return("income.net_income", "balance.total_assets"),
            _FORMAT_PERCENT,
            "Net income ÷ average total assets",
        ),
        _Derived(
            "Current ratio",
            ratio("balance.current_assets", "balance.current_liabilities"),
            _FORMAT_MULTIPLE,
            "Current assets ÷ current liabilities",
        ),
        _Derived(
            "Quick ratio",
            quick_ratio,
            _FORMAT_MULTIPLE,
            "(Current assets − inventory) ÷ current liabilities",
        ),
        _Derived(
            "Debt to equity",
            debt_to_equity,
            _FORMAT_MULTIPLE,
            "(Short-term debt + long-term debt) ÷ total equity",
        ),
        _Derived(
            "Free cash flow",
            free_cash_flow,
            _FORMAT_CURRENCY,
            "Operating cash flow − capital expenditure",
        ),
        _Derived(
            "Free cash flow conversion",
            fcf_conversion,
            _FORMAT_PERCENT,
            "Free cash flow ÷ net income",
        ),
        _Derived(
            "Days sales outstanding",
            days_sales_outstanding,
            _FORMAT_DAYS,
            "Accounts receivable ÷ revenue × days in year",
        ),
        _Derived(
            "Days inventory",
            days_inventory,
            _FORMAT_DAYS,
            "Inventory ÷ cost of revenue × days in year",
        ),
        _Derived(
            "Revenue growth",
            growth("income.revenue"),
            _FORMAT_PERCENT,
            "Revenue ÷ prior-year revenue − 1",
        ),
        _Derived(
            "Operating income growth",
            growth("income.operating_income"),
            _FORMAT_PERCENT,
            "Operating income ÷ prior year − 1",
        ),
        _Derived(
            "Net income growth",
            growth("income.net_income"),
            _FORMAT_PERCENT,
            "Net income ÷ prior year − 1",
        ),
    )


def assemble_xlsx(document: ReportDocument) -> bytes:
    """Renders the underlying figures as a workbook.

    Raises:
        ArtifactRenderError: openpyxl is unavailable, or the workbook could not
            be written. Never raised at a reader — the caller reports it.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, Side
        from openpyxl.utils import get_column_letter
    except ImportError as cause:
        message = (
            "The spreadsheet renderer is not available on this server, so "
            "this report was not published as a workbook."
        )
        raise ArtifactRenderError(message) from cause

    import io

    facts = {fact.id: fact for fact in document.facts}
    index = _build_index(list(document.facts))
    years = index.years

    workbook = Workbook()
    # The default sheet becomes the cover rather than being deleted and
    # recreated, which keeps it first in the tab order.
    cover = workbook.active
    if cover is None:  # pragma: no cover — a new Workbook always has one
        cover = workbook.create_sheet()
    cover.title = _SHEET_COVER

    styles = _WorkbookStyles(
        heading=Font(name="Calibri", size=14, bold=True, color="FF14201C"),
        label=Font(name="Calibri", size=10, color="FF14201C"),
        muted=Font(name="Calibri", size=9, color="FF6C7671"),
        column=Font(name="Consolas", size=9, bold=True, color="FF1F4D3D"),
        figure=Font(name="Consolas", size=10, color="FF14201C"),
        total=Font(name="Consolas", size=10, bold=True, color="FF14201C"),
        rule=Border(bottom=Side(style="thin", color="FFE2DED4")),
        heavy=Border(bottom=Side(style="medium", color="FF14201C")),
        right=Alignment(horizontal="right"),
        wrap=Alignment(horizontal="left", vertical="top", wrap_text=True),
        column_letter=get_column_letter,
    )

    _write_cover(cover, document, styles)

    cells: dict[str, str] = {}
    for title, rows in (
        (_SHEET_INCOME, _INCOME_ROWS),
        (_SHEET_BALANCE, _BALANCE_ROWS),
        (_SHEET_CASHFLOW, _CASHFLOW_ROWS),
    ):
        sheet = workbook.create_sheet(title)
        _write_statement(sheet, title, rows, index, years, styles, cells)

    _write_analysis(workbook.create_sheet(_SHEET_ANALYSIS), years, cells, styles)

    if index.segments:
        _write_segments(
            workbook.create_sheet(_SHEET_SEGMENTS), index, years, styles
        )

    _write_assumptions(
        workbook.create_sheet(_SHEET_ASSUMPTIONS), document, styles
    )
    _write_sources(
        workbook.create_sheet(_SHEET_SOURCES), document, facts, styles
    )

    buffer = io.BytesIO()
    try:
        workbook.save(buffer)
    except Exception as cause:  # noqa: BLE001 — surfaced as a typed failure
        logger.error("Workbook could not be written", exc_info=cause)
        message = "This report could not be rendered as a workbook."
        raise ArtifactRenderError(message) from cause

    return buffer.getvalue()


@dataclass(frozen=True, slots=True)
class _WorkbookStyles:
    """Fonts, borders and alignments, resolved once.

    Held in a value rather than reached for globally so that importing this
    module never requires openpyxl — only rendering does.
    """

    heading: Font
    label: Font
    muted: Font
    column: Font
    figure: Font
    total: Font
    rule: Border
    heavy: Border
    right: Alignment
    wrap: Alignment
    column_letter: Callable[[int], str]


def _write_cover(
    sheet: Worksheet, document: ReportDocument, styles: _WorkbookStyles
) -> None:
    """Identity, provenance summary and what the other tabs contain."""
    company = document.company
    compliance = document.compliance
    completed = document.report.completed_at or document.report.created_at

    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 74

    sheet["A1"] = company.name
    sheet["A1"].font = styles.heading
    sheet["A2"] = "Company profile, sourced from SEC filings"
    sheet["A2"].font = styles.muted

    entries: tuple[tuple[str, _CellValue], ...] = (
        ("Ticker", company.ticker),
        ("CIK", company.cik),
        ("Filer type", _filer_form(company)),
        ("SIC code", company.sic_code or NOT_DISCLOSED_TEXT),
        ("Sector", company.sector or NOT_DISCLOSED_TEXT),
        ("Reporting currency", company.reporting_currency or NOT_DISCLOSED_TEXT),
        ("Fiscal year end", company.fiscal_year_end or NOT_DISCLOSED_TEXT),
        ("Report id", document.report.id),
        ("Generated", completed.isoformat(timespec="seconds")),
        ("", ""),
        ("Facts stored", compliance.fact_count),
        ("Figures in prose", compliance.figure_count),
        ("Figures cited", compliance.cited_figure_count),
        ("Citation coverage", compliance.coverage_display),
        (
            "Verification",
            "Passed" if compliance.passed else "Blocking issues found",
        ),
        ("Verified at", compliance.verified_at.isoformat(timespec="seconds")),
        ("", ""),
        (
            "How to read this workbook",
            "Statement tabs hold reported figures only. The analysis tab "
            "holds no values — every cell is a formula over those figures, "
            "so changing an input changes the result. The assumptions tab "
            "holds the tunables those formulas reference. Every figure is "
            "listed on the sources tab with the filing it came from.",
        ),
    )

    row = 4
    for label, value in entries:
        sheet.cell(row=row, column=1, value=label).font = styles.label
        cell = sheet.cell(row=row, column=2, value=value)
        cell.font = styles.figure if label else styles.muted
        if label == "How to read this workbook":
            cell.alignment = styles.wrap
            sheet.row_dimensions[row].height = 74
        row += 1


def _write_statement(
    sheet: Worksheet,
    title: str,
    rows: Sequence[_Row],
    index: _Index,
    years: Sequence[int],
    styles: _WorkbookStyles,
    cells: dict[str, str],
) -> None:
    """One statement tab: reported figures, one column per fiscal year.

    Records every written cell in `cells` so the analysis tab can point at it.
    That map is the whole reason the workbook's formulas are live rather than
    values pasted in from Python.
    """
    sheet.column_dimensions["A"].width = 38
    sheet["A1"] = title
    sheet["A1"].font = styles.heading
    sheet["A2"] = (
        "Reported figures, as filed. Nothing on this tab is calculated."
    )
    sheet["A2"].font = styles.muted

    header_row = 4
    sheet.cell(row=header_row, column=1, value="Figure").font = styles.column
    sheet.cell(row=header_row, column=1).border = styles.heavy

    for offset, year in enumerate(years):
        column = 2 + offset
        cell = sheet.cell(row=header_row, column=column, value=_period_label(year))
        cell.font = styles.column
        cell.alignment = styles.right
        cell.border = styles.heavy
        sheet.column_dimensions[styles.column_letter(column)].width = 18

    row_at = header_row + 1
    for row in rows:
        written = False
        for offset, year in enumerate(years):
            fact = index.fact(row.metric, year)
            if fact is None or fact.value is None:
                continue
            column = 2 + offset
            cell = sheet.cell(row=row_at, column=column, value=fact.value)
            cell.font = (
                styles.total
                if row.emphasis is FigureEmphasis.TOTAL
                else styles.figure
            )
            cell.number_format = (
                _FORMAT_PER_SHARE
                if row.metric in _PER_SHARE_METRICS
                else _FORMAT_CURRENCY
            )
            cell.border = styles.rule
            cells[f"{row.metric}@{year}"] = (
                f"'{title}'!{styles.column_letter(column)}{row_at}"
            )
            written = True

        if not written:
            continue

        label = sheet.cell(row=row_at, column=1, value=row.label)
        label.font = styles.label
        label.border = styles.rule
        row_at += 1

    sheet.freeze_panes = "B5"


def _write_analysis(
    sheet: Worksheet,
    years: Sequence[int],
    cells: Mapping[str, str],
    styles: _WorkbookStyles,
) -> None:
    """Derived metrics, as formulas over the statement tabs.

    No value on this tab is written by Python. Every cell is an Excel formula
    referencing the reported figures, which is what makes the workbook
    auditable rather than a second copy of the report.
    """
    sheet.column_dimensions["A"].width = 32
    sheet["A1"] = "Analysis"
    sheet["A1"].font = styles.heading
    sheet["A2"] = (
        "Every cell below is a formula over the statement tabs. Change a "
        "reported figure and these move with it."
    )
    sheet["A2"].font = styles.muted

    header_row = 4
    sheet.cell(row=header_row, column=1, value="Metric").font = styles.column
    sheet.cell(row=header_row, column=1).border = styles.heavy

    for offset, year in enumerate(years):
        column = 2 + offset
        cell = sheet.cell(
            row=header_row, column=column, value=_period_label(year)
        )
        cell.font = styles.column
        cell.alignment = styles.right
        cell.border = styles.heavy
        sheet.column_dimensions[styles.column_letter(column)].width = 18

    formula_column = 2 + len(years)
    note_cell = sheet.cell(row=header_row, column=formula_column, value="Formula")
    note_cell.font = styles.column
    note_cell.border = styles.heavy
    sheet.column_dimensions[styles.column_letter(formula_column)].width = 46

    row_at = header_row + 1
    for derived in _formula_rows():
        written = False
        for offset, year in enumerate(years):
            resolver = _resolver(cells, year)
            formula = derived.build(resolver)
            if formula is None:
                continue
            cell = sheet.cell(row=row_at, column=2 + offset, value=formula)
            cell.font = styles.figure
            cell.number_format = derived.number_format
            cell.border = styles.rule
            written = True

        if not written:
            continue

        label = sheet.cell(row=row_at, column=1, value=derived.label)
        label.font = styles.label
        label.border = styles.rule
        note = sheet.cell(row=row_at, column=formula_column, value=derived.note)
        note.font = styles.muted
        note.border = styles.rule
        row_at += 1

    sheet.freeze_panes = "B5"


def _resolver(cells: Mapping[str, str], year: int) -> _CellRef:
    """Resolves a metric reference to a cell for one period.

    "income.revenue" is this year's cell; "income.revenue@-1" is last year's.
    Returns None when that figure was not reported, which is how a formula
    declines to exist rather than dividing by a blank.
    """

    def at(reference: str) -> str | None:
        if reference.endswith("@-1"):
            return cells.get(f"{reference[:-3]}@{year - 1}")
        return cells.get(f"{reference}@{year}")

    return at


def _write_segments(
    sheet: Worksheet,
    index: _Index,
    years: Sequence[int],
    styles: _WorkbookStyles,
) -> None:
    """Revenue by segment, as the filer disclosed it."""
    sheet.column_dimensions["A"].width = 38
    sheet["A1"] = "Revenue by segment"
    sheet["A1"].font = styles.heading
    sheet["A2"] = (
        "Dimensional figures read from the filing's XBRL instance document."
    )
    sheet["A2"].font = styles.muted

    header_row = 4
    sheet.cell(row=header_row, column=1, value="Segment").font = styles.column
    sheet.cell(row=header_row, column=1).border = styles.heavy
    for offset, year in enumerate(years):
        column = 2 + offset
        cell = sheet.cell(
            row=header_row, column=column, value=_period_label(year)
        )
        cell.font = styles.column
        cell.alignment = styles.right
        cell.border = styles.heavy
        sheet.column_dimensions[styles.column_letter(column)].width = 18

    row_at = header_row + 1
    for label, by_year in index.segments.items():
        wrote = False
        for offset, year in enumerate(years):
            fact = by_year.get(year)
            if fact is None or fact.value is None:
                continue
            cell = sheet.cell(row=row_at, column=2 + offset, value=fact.value)
            cell.font = styles.figure
            cell.number_format = _FORMAT_CURRENCY
            cell.border = styles.rule
            wrote = True
        if not wrote:
            continue
        name = sheet.cell(row=row_at, column=1, value=label)
        name.font = styles.label
        name.border = styles.rule
        row_at += 1

    sheet.freeze_panes = "B5"


def _write_assumptions(
    sheet: Worksheet, document: ReportDocument, styles: _WorkbookStyles
) -> None:
    """The tunables, split into the ones formulas use and the ones they don't.

    The first block is referenced by the analysis tab — change days in year and
    the working-capital rows move. The second block records what was assumed
    during extraction, so a reader can see the rules a figure was gathered
    under even though no formula reads them.
    """
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 70

    sheet["A1"] = "Assumptions"
    sheet["A1"].font = styles.heading

    sheet["A3"] = "Used in formulas"
    sheet["A3"].font = styles.column
    sheet["A3"].border = styles.heavy
    sheet["B3"] = "Value"
    sheet["B3"].font = styles.column
    sheet["B3"].border = styles.heavy
    sheet["C3"] = "Why it is what it is"
    sheet["C3"].font = styles.column
    sheet["C3"].border = styles.heavy

    # Row 4 is referenced by _ASSUMPTION_DAYS_CELL. Moving it means changing
    # that constant, which is why it is pinned here rather than appended.
    sheet["A4"] = "Days in year"
    sheet["A4"].font = styles.label
    sheet["B4"] = DAYS_IN_YEAR
    sheet["B4"].font = styles.figure
    sheet["C4"] = (
        "The calendar year, not the fiscal one. A 53-week filer's cycle is "
        "still quoted in ordinary days, and mixing the two would make years "
        "incomparable."
    )
    sheet["C4"].font = styles.muted
    sheet["C4"].alignment = styles.wrap
    sheet.row_dimensions[4].height = 44

    documented: tuple[tuple[str, _CellValue, str], ...] = (
        (
            "Periods tabulated",
            len(document.charts.revenue_margin.points)
            if document.charts.revenue_margin
            else 0,
            "Annual periods shown, most recent last. Set by the requested "
            "analysis depth.",
        ),
        (
            "Reporting currency",
            document.company.reporting_currency or NOT_DISCLOSED_TEXT,
            "Read from the filer's own XBRL unit keys, never assumed to be "
            "US dollars.",
        ),
        (
            "Annual period bounds",
            "300–400 days",
            "A 52/53-week fiscal calendar makes 'a year' anything from 358 "
            "to 371 days, so the window is wide enough to admit one and "
            "narrow enough to exclude a half-year.",
        ),
        (
            "Amendment precedence",
            "Amendments supersede originals",
            "For the same period, a 10-K/A is preferred over the 10-K it "
            "amends.",
        ),
        (
            "Segment sum tolerance",
            "0.5%",
            "Filings round segment tables independently of the income "
            "statement, so segments may miss the consolidated total by this "
            "fraction before it is reported as a discrepancy.",
        ),
        (
            "Balance sheet tolerance",
            "0.5%",
            "Assets minus liabilities minus equity, as a fraction of assets.",
        ),
        (
            "Cash flow tie tolerance",
            "2%",
            "Wider than the balance sheet's on purpose: the gap is the "
            "effect of exchange rates on cash, which is reported on its own "
            "line and is not extracted as a metric.",
        ),
        (
            "Required citation coverage",
            "100%",
            "Every figure in the report must resolve to a stored fact. "
            "'Most figures are sourced' is not the product.",
        ),
        (
            "Market data tier",
            "Tier 3",
            "Price and market capitalisation only, and hard-blocked from the "
            "financial highlights in code rather than by convention.",
        ),
    )

    row = 7
    sheet.cell(row=6, column=1, value="Applied during extraction").font = (
        styles.column
    )
    sheet.cell(row=6, column=1).border = styles.heavy
    sheet.cell(row=6, column=2, value="Value").font = styles.column
    sheet.cell(row=6, column=2).border = styles.heavy
    sheet.cell(row=6, column=3, value="Why it is what it is").font = styles.column
    sheet.cell(row=6, column=3).border = styles.heavy

    for label, value, why in documented:
        sheet.cell(row=row, column=1, value=label).font = styles.label
        sheet.cell(row=row, column=2, value=value).font = styles.figure
        note = sheet.cell(row=row, column=3, value=why)
        note.font = styles.muted
        note.alignment = styles.wrap
        sheet.row_dimensions[row].height = 42
        row += 1


def _write_sources(
    sheet: Worksheet,
    document: ReportDocument,
    facts: Mapping[str, DocumentFact],
    styles: _WorkbookStyles,
) -> None:
    """Every fact in the report, with the filing behind it.

    This tab is why the workbook can be handed to someone who does not trust
    it: each figure names its metric, its period, its tier, the accession
    number it came from, the XBRL tag that resolved, and — where the figure was
    derived — the formula that produced it.
    """
    columns: tuple[tuple[str, int], ...] = (
        ("Metric", 34),
        ("Label", 34),
        ("Value", 18),
        ("Displayed as", 18),
        ("Unit", 10),
        ("Period start", 13),
        ("Period end", 13),
        ("Fiscal year", 11),
        ("Segment", 24),
        ("Tier", 6),
        ("Source type", 16),
        ("Form", 10),
        ("Accession", 22),
        ("Filed", 12),
        ("Extraction", 20),
        ("Confidence", 11),
        ("XBRL tag", 40),
        ("Taxonomy", 12),
        ("Formula", 52),
        ("Document", 72),
    )

    sheet["A1"] = "Sources"
    sheet["A1"].font = styles.heading
    sheet["A2"] = (
        "One row per fact. A figure that cannot name its source was never "
        "stored, so there are no blanks in the provenance columns."
    )
    sheet["A2"].font = styles.muted

    header_row = 4
    for position, (title, width) in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=position, value=title)
        cell.font = styles.column
        cell.border = styles.heavy
        sheet.column_dimensions[styles.column_letter(position)].width = width

    forms = {filing.accession_no: filing.form for filing in document.filings}

    row_at = header_row + 1
    for fact in facts.values():
        values: tuple[_CellValue, ...] = (
            fact.metric,
            fact.label,
            fact.value,
            fact.display_value,
            fact.unit or "",
            fact.period_start.isoformat() if fact.period_start else "",
            fact.period_end.isoformat(),
            fact.fiscal_year or "",
            fact.segment_label or fact.segment_member or "",
            int(fact.tier),
            str(fact.source_type),
            forms.get(fact.accession_no, ""),
            fact.accession_no,
            fact.filed_date.isoformat(),
            str(fact.extraction_method),
            fact.confidence,
            fact.resolved_tag or "",
            str(fact.taxonomy) if fact.taxonomy else "",
            fact.formula or "",
            str(fact.source_url),
        )
        for position, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_at, column=position, value=value)
            cell.font = styles.figure if position == 3 else styles.label
            cell.border = styles.rule
            if position == 3 and isinstance(value, float):
                cell.number_format = _FORMAT_CURRENCY
        row_at += 1

    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = (
        f"A{header_row}:{styles.column_letter(len(columns))}{row_at - 1}"
    )


# --- Publishing --------------------------------------------------------------


async def publish(
    report_id: str, document: ReportDocument
) -> tuple[ArtifactRef, ...]:
    """Renders both artifacts and uploads them.

    Neither rendering nor uploading can fail the report: an artifact that could
    not be produced comes back carrying the reason, which the interface shows
    in place of a download link.
    """
    refs: list[ArtifactRef] = []

    for kind, render in (
        (ArtifactKind.PDF, assemble_pdf),
        (ArtifactKind.XLSX, assemble_xlsx),
    ):
        try:
            content = render(document)
        except Exception as failure:  # noqa: BLE001 — see below
            # Broad on purpose. A renderer is a leaf of the pipeline and the
            # report is already verified by the time it runs, so no failure
            # here — typed or not — may take the report down with it. A
            # `ArtifactRenderError` carries a message written for a reader;
            # anything else gets one, and the detail goes to the log.
            reason = (
                str(failure)
                if isinstance(failure, ArtifactRenderError)
                else (
                    f"This report could not be rendered as {str(kind).upper()}."
                )
            )
            logger.warning(
                "Artifact could not be rendered",
                extra={
                    "report_id": report_id,
                    "kind": str(kind),
                    "error": str(failure),
                },
                exc_info=not isinstance(failure, ArtifactRenderError),
            )
            refs.append(
                ArtifactRef(
                    kind=kind,
                    size_bytes=0,
                    content_type=ARTIFACT_CONTENT_TYPES[str(kind)],
                    created_at=dt.datetime.now(dt.UTC),
                    unavailable_reason=reason,
                )
            )
            continue

        refs.append(
            await storage.upload(
                report_id, document.company.ticker, kind, content
            )
        )

    return tuple(refs)
